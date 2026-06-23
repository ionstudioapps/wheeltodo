"""Build the WheelTodo 30-day beta-launch content calendar (.xlsx).
Grounded in real shipped features only. Body-double = roadmap tease (Day 30).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Brand palette (from packages/shared/src/themes.ts)
CORAL = "FF5C4D"
INK = "111111"
BEIGE = "E8E0D5"
BEIGE_LT = "F5F1EB"
WHITE = "FFFFFF"
PURPLE = "A78BFA"   # rest accent used in-app
TEAL = "2BB6A3"

FONT = "Arial"
thin = Side(style="thin", color="D8D0C4")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ----------------------------------------------------------------------------
# Sheet 1 — 30-Day Calendar
# ----------------------------------------------------------------------------
cal = wb.active
cal.title = "30-Day Calendar"

headers = ["Day", "Weekday", "Phase", "Platform", "Pillar", "Format",
           "Concept / Working title", "Visual hook (0-1s)", "Verbal hook",
           "Text overlay", "CTA", "Asset needed", "Status"]

# Phase color map
PHASE_COLOR = {
    "Pre-launch": "FCE4DF",
    "LAUNCH WEEK": CORAL,
    "Use-cases": "EDE7DC",
    "Retention": "E4F3EF",
}

TT = "TT · IG Reels · Shorts"
STORY = "Story / static only"

# day, weekday, phase, platform, pillar, format, concept, visual, verbal, overlay, cta, asset, status
rows = [
 (1,"Mon","Pre-launch",TT,"Relatable pain","Talking head","'23 tasks, 0 done' — name the freeze","Overflowing list, phone face-down, staring at ceiling","I wasn't lazy. I was paralyzed.","23 tasks. 0 done. there's a reason.","Follow — the fix drops this week","TH clip + list mockup","Planned"),
 (2,"Tue","Pre-launch",TT,"Niche · ADHD","POV / text","'Your brain has 40 tabs open'","Frantic tab-switching metaphor","ADHD brains can't pick a starting point.","why 'just start' never works for us","Comment if this is you","Screen-rec + TH","Planned"),
 (3,"Wed","Pre-launch",TT,"Oddly satisfying","Teaser b-roll","First peek at the wheel (no name reveal)","Full-screen wheel mid-spin, crisp tick, lands on one","(whispered) just pick one for me…","what if you never had to choose again 👀","Waitlist link in bio","Spin screen-rec","Planned"),
 (4,"Thu","Pre-launch",TT,"Relatable pain","Skit","The procrastination spiral","'Getting ready to be productive' — watering plant, tidying desk","Me 'about to start' for the 4th time today","POV: you've 'started' 6 times today","Follow for the actual fix","Skit","Planned"),
 (5,"Fri","Pre-launch",TT,"Trend-jack","Green-screen","React to 'just make a list' advice","Green-screen over a productivity influencer","The list IS the problem — I can't pick.","telling an overwhelmed person to 'make a list' 💀","Real fix drops Monday","Green-screen edit","Planned"),
 (6,"Sat","Pre-launch",TT,"Oddly satisfying","ASMR","Spin compilation (loopable)","3–4 spins back-to-back, tick + land","(none / whispered)","the most satisfying way to start a task","Name drops next week 👀","Spin compilation","Planned"),
 (7,"Sun","Pre-launch",STORY,"Rest day (on-brand)","Story only","We rest. So does our app. 🌿","Calm static — streak flame staying lit on a rest day","—","we take rest days. soon you'll see why that matters.","—","Static Rest Mode card","Planned"),
 (8,"Mon","LAUNCH WEEK",TT,"Relatable pain","Hero reveal","THE reveal — core loop","Dump list → spin → lands → instantly flips to a RUNNING Pomodoro","Stop choosing. Spin — the timer's already counting.","meet WheelTodo · 23 tasks · one spin · now you're already doing it","Beta's live — link in bio","Hero screen-rec","Planned"),
 (9,"Tue","LAUNCH WEEK",TT,"Differentiator: kind+rest","Emotional","'Rest counts too'","Collapse on couch, wrecked; phone shows streak flame STAYING lit","I took a nap and my to-do app… congratulated me.","the productivity app that rewards you for resting","Join the beta","Couch clip + streak UI","Planned"),
 (10,"Wed","LAUNCH WEEK",TT,"Niche · ADHD","Feature demo","The energy check-in","App asks 'How's your energy?' → taps Drained → suggests gentle wins","It asks how I feel before it asks what I'll do.","a to-do app that checks on you first","Beta link in bio","Check-in screen-rec","Planned"),
 (11,"Thu","LAUNCH WEEK",TT,"Proof & momentum","Vlog","'I let a wheel run my whole day' — Day 1 · SERIES","Morning: 'the wheel decides everything today'","For 24h I'm not allowed to choose a single task.","I let an app control my entire day","Follow for Day 2","Day-in-life montage","Planned"),
 (12,"Fri","LAUNCH WEEK",TT,"Oddly satisfying","ASMR · sound-on","Spin + tick + completion confetti","Spin → tick → lands → confetti on complete","(none)","turn your sound on 🔊","Beta link in bio","Spin+confetti (gen-audio SFX)","Planned"),
 (13,"Sat","LAUNCH WEEK",TT,"Proof & momentum","Testimonial","First beta-tester reactions","Screenshot carousel of DMs / reactions","—","what beta testers are saying after 3 days","Join them","Screenshot reel","Planned"),
 (14,"Sun","LAUNCH WEEK",STORY,"Rest day (on-brand)","Story only","Sunday = rest day, streak safe","Calm static","—","Sunday. your streak's safe. 🌿","—","Static card","Planned"),
 (15,"Mon","Use-cases",TT,"Niche · students","Study-with-me","'The app picks what to study'","Desk study clip, spin selects the subject","study with me — but it makes the decision","study with me (the app picks)","Beta link in bio","Desk study clip","Planned"),
 (16,"Tue","Use-cases",TT,"Niche · couples · FRANCHISE","Skit","'Who does the dishes?'","Two people glaring at a full sink, phone between them","We argue about chores daily. Not anymore.","we let the wheel decide who does the dishes","Tag your roommate","Skit","Planned"),
 (17,"Wed","Use-cases",TT,"Differentiator: kind+rest","Educational","'Streaks shouldn't punish rest'","Explainer: rest protects the streak","Every streak app breaks if you rest. Ours doesn't.","rest days keep your streak alive","Join the beta","UI explainer","Planned"),
 (18,"Thu","Use-cases",TT,"Proof & momentum","Vlog","Wheel-runs-my-day — Day 2 / payoff · SERIES","Results montage","Most productive day in weeks?","I let a wheel run my day — here's what happened","Follow","Montage","Planned"),
 (19,"Fri","Use-cases",TT,"Trend-jack","'Apps that replaced X'","'Deleted 3 to-do apps for this'","Dramatically deleting Notion / Reminders / sticky notes","I never DID the to-dos. This one makes me.","the to-do app that actually makes you do it","Beta link in bio","Phone screen","Planned"),
 (20,"Sat","Use-cases",TT,"Niche · overwhelm","Relatable","'When everything needs doing and you freeze'","Quiet overwhelm, too many sticky notes","you're not lazy — you're overloaded","when everything's urgent and you freeze","Beta link in bio","Talking head","Planned"),
 (21,"Sun","Use-cases",STORY,"Rest day (on-brand)","Story only","Community poll","Static poll sticker","—","what's your Sunday rest activity? 🌿","Reply to vote","Poll story","Planned"),
 (22,"Mon","Retention",TT,"Oddly satisfying","ASMR","Theme aesthetics (4 themes)","Theme switch screen-rec, incl. dark","(none)","pick your vibe 🎨 (dark mode included)","Beta link in bio","Theme switch rec","Planned"),
 (23,"Tue","Retention",TT,"Differentiator: kind+rest","Values","Accessible high-contrast themes","A11y theme demo side-by-side","built for every brain — and every pair of eyes","accessible by default, not as an afterthought","Beta link in bio","A11y demo","Planned"),
 (24,"Wed","Retention",TT,"Niche · ADHD","Educational","'Task initiation is the hard part'","Talking head + UI","The hardest part of any task is choosing it. So we deleted choosing.","task paralysis, solved","Join the beta","TH + UI","Planned"),
 (25,"Thu","Retention",TT,"Proof & momentum","Duet/stitch prompt","'Show me your wheel' challenge","Prompt clip inviting stitches","—","stitch this with your task wheel","Use #WheelTodo","Prompt clip","Planned"),
 (26,"Fri","Retention",TT,"Trend-jack","Trending audio","'Living in my head rent-free'","Meme edit: 17 unstarted tasks → one spin","(trending audio)","my 17 unstarted tasks vs. one spin","Beta link in bio","Meme edit","Planned"),
 (27,"Sat","Retention",TT,"Differentiator: kind+rest","Streak flex","'14-day streak — rested 4 of them'","Shareable streak card","guilt-free productivity is real","14-day streak (and I rested 4 days)","Beta link in bio","Streak share card","Planned"),
 (28,"Sun","Retention",STORY,"Rest day (on-brand)","Story only","Week recap","Calm recap static","—","another week, streak intact. 🌿","—","Recap story","Planned"),
 (29,"Mon","Retention",TT,"Proof & momentum","Montage","'3 weeks of beta — what changed'","Before/after montage","from 23-tasks-frozen to actually done","3 weeks in: what changed","Join the beta","Before/after montage","Planned"),
 (30,"Tue","Retention",TT,"Proof & momentum","Teaser","Body-double companion (ROADMAP — not yet shipped)","Cozy lofi room loop preview (mark as 'soon')","a little companion that sits with you while you work…","your focus, but make it lofi 🌧️ (soon)","Join beta to get it first","body-double loop (repo asset · unreleased)","Planned"),
]

# title row
cal.merge_cells("A1:M1")
t = cal["A1"]
t.value = "WheelTodo — 30-Day Beta-Launch Content Calendar"
t.font = Font(name=FONT, size=15, bold=True, color=WHITE)
t.fill = PatternFill("solid", fgColor=INK)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
cal.row_dimensions[1].height = 30

sub = cal["A2"]
cal.merge_cells("A2:M2")
sub.value = ("Day 1 = first Monday of the push (set real dates in your tool). 26 posting days + 4 on-brand rest days.  "
             "Cross-post all reels to TikTok / IG Reels / YouTube Shorts.  Caption the verbal hook (most watch muted).")
sub.font = Font(name=FONT, size=9, italic=True, color="6B6258")
sub.fill = PatternFill("solid", fgColor=BEIGE_LT)
sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
cal.row_dimensions[2].height = 26

HEADER_ROW = 3
for c, h in enumerate(headers, 1):
    cell = cal.cell(row=HEADER_ROW, column=c, value=h)
    cell.font = Font(name=FONT, size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=CORAL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
cal.row_dimensions[HEADER_ROW].height = 28

r = HEADER_ROW + 1
for row in rows:
    is_rest = "Rest day" in row[4]
    phase = row[2]
    accent = PHASE_COLOR.get(phase, WHITE)
    for c, val in enumerate(row, 1):
        cell = cal.cell(row=r, column=c, value=val)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.font = Font(name=FONT, size=9, color=INK)
        if c == 1:  # Day
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name=FONT, size=11, bold=True, color=CORAL if not is_rest else "9A8FB0")
        if c == 3:  # Phase column gets the accent chip
            launch = phase == "LAUNCH WEEK"
            cell.fill = PatternFill("solid", fgColor=accent)
            cell.font = Font(name=FONT, size=9, bold=True, color=WHITE if launch else INK)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if is_rest:
        for c in range(1, 14):
            if c != 3:
                cal.cell(row=r, column=c).fill = PatternFill("solid", fgColor="F3F0FA")
    elif (r % 2) == 0:
        for c in range(1, 14):
            if c != 3:
                cal.cell(row=r, column=c).fill = PatternFill("solid", fgColor=BEIGE_LT)
    r += 1

widths = [5,9,12,15,16,14,26,30,28,28,20,22,10]
for i, w in enumerate(widths, 1):
    cal.column_dimensions[get_column_letter(i)].width = w
cal.freeze_panes = "A4"

# ----------------------------------------------------------------------------
# Sheet 2 — Pillars
# ----------------------------------------------------------------------------
pil = wb.create_sheet("Pillars")
pil.merge_cells("A1:E1")
pt = pil["A1"]; pt.value = "Content Pillars (repeatable, not 10 one-offs)"
pt.font = Font(name=FONT, size=14, bold=True, color=WHITE); pt.fill = PatternFill("solid", fgColor=INK)
pt.alignment = Alignment(horizontal="left", vertical="center", indent=1); pil.row_dimensions[1].height = 28

ph = ["Pillar", "Why it works", "Target mix", "Core message", "Best formats"]
for c, h in enumerate(ph, 1):
    cell = pil.cell(row=2, column=c, value=h)
    cell.font = Font(name=FONT, size=10, bold=True, color=WHITE); cell.fill = PatternFill("solid", fgColor=CORAL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = border
pil.row_dimensions[2].height = 24

pillars = [
 ("Relatable pain","Widest top-of-funnel reach — decision paralysis / overwhelm is instantly recognisable. Front-loaded pre-launch.","3 / 26","You're not lazy, you're frozen — and that's fixable","Talking head, skit, POV text"),
 ("Differentiator: kind+rest","The ownable wedge no competitor holds — rest protects your streak + an accessible, kind brand. Emotional + shareable.","4 / 26","Rest counts. Guilt-free, built for every brain.","Emotional, educational, UI demo, values"),
 ("Oddly satisfying","Dopamine bait — the spin/tick loops; high watch-time + replays","4 / 26","The spin is the payoff","ASMR, loopable b-roll, streak flex"),
 ("Niche use-cases","High save/share — ADHD, students, couples, parents","6 / 26","Made for your specific brain / situation","Feature demo, skit, study-with-me"),
 ("Trend-jack","Rides reach of existing trends, formats & audio","3 / 26","We belong in this conversation","Green-screen, 'apps that replaced X', trending audio"),
 ("Proof & momentum","Credibility + community — testimonials, the wheel-runs-my-day series, UGC, roadmap teases. Back-loaded as the beta matures and proof accumulates.","6 / 26","Real people, real momentum — and what's next","Testimonial, vlog/series, UGC prompt, teaser"),
]
PIL_NOTE = ("Shape over time matters more than the raw mix: days 1–10 skew pain + satisfying (reach), days 11–30 add "
            "proof + niche (save/share + conversion). Plus 4 on-brand rest days = no reel, story only.")
rr = 3
for p in pillars:
    for c, val in enumerate(p, 1):
        cell = pil.cell(row=rr, column=c, value=val); cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.font = Font(name=FONT, size=9, color=INK)
        if c == 1: cell.font = Font(name=FONT, size=10, bold=True, color=CORAL)
        if c == 3: cell.alignment = Alignment(horizontal="center", vertical="center")
    rr += 1
pil.merge_cells(start_row=rr+1, start_column=1, end_row=rr+1, end_column=5)
nt = pil.cell(row=rr+1, column=1, value=PIL_NOTE)
nt.font = Font(name=FONT, size=9, italic=True, color="6B6258")
nt.fill = PatternFill("solid", fgColor=BEIGE_LT)
nt.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
pil.row_dimensions[rr+1].height = 30
for col, w in zip("ABCDE", [24,46,9,34,30]):
    pil.column_dimensions[col].width = w

# ----------------------------------------------------------------------------
# Sheet 3 — Hook Bank
# ----------------------------------------------------------------------------
hb = wb.create_sheet("Hook Bank")
hb.merge_cells("A1:C1")
ht = hb["A1"]; ht.value = "Hook Bank — mix & match across any reel (first 1s is everything)"
ht.font = Font(name=FONT, size=14, bold=True, color=WHITE); ht.fill = PatternFill("solid", fgColor=INK)
ht.alignment = Alignment(horizontal="left", vertical="center", indent=1); hb.row_dimensions[1].height = 28

for c, h in enumerate(["Visual hooks (frame 0–1s)", "Verbal hooks (first line)", "Text-overlay hooks (burned-in)"], 1):
    cell = hb.cell(row=2, column=c, value=h)
    cell.font = Font(name=FONT, size=10, bold=True, color=WHITE); cell.fill = PatternFill("solid", fgColor=CORAL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = border
hb.row_dimensions[2].height = 24

visual = ["Full-screen wheel frozen mid-spin","Overflowing chaotic list, phone face-down","You frozen, blank 'deer-in-headlights' stare",
          "Satisfying tick as wheel lands on one task","Spin → screen instantly flips to a running timer","Collapse on couch; streak flame stays lit",
          "Before/after: 23 unchecked boxes → 1 checked","Energy check-in: tap 'Drained' → gentle tasks appear"]
verbal = ["I wasn't lazy — I was paralyzed.","I let an app decide my entire day.","If your brain has 40 tabs open, watch this.",
          "The hardest part of any task is choosing it. So we deleted choosing.","I took a nap and my to-do app congratulated me.",
          "Stop choosing. Spin — the timer's already counting.","Every streak app breaks if you rest. Ours doesn't.","It asks how I feel before it asks what I'll do."]
overlay = ["23 tasks. 0 done. here's the fix →","POV: you've 'started' 6 times today","don't choose your task. spin it.",
           "the to-do app that actually makes you do it","the productivity app that rewards you for resting","a to-do app that checks on you first",
           "rest days keep your streak alive","the most satisfying way to start working"]
maxn = max(len(visual), len(verbal), len(overlay))
for i in range(maxn):
    for c, src in enumerate([visual, verbal, overlay], 1):
        val = src[i] if i < len(src) else ""
        cell = hb.cell(row=3+i, column=c, value=val); cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.font = Font(name=FONT, size=9, color=INK)
        if (i % 2) == 1:
            cell.fill = PatternFill("solid", fgColor=BEIGE_LT)
for col in "ABC":
    hb.column_dimensions[col].width = 40

# ----------------------------------------------------------------------------
# Sheet 4 — Cadence & Channels
# ----------------------------------------------------------------------------
cc = wb.create_sheet("Cadence & Channels")
cc.merge_cells("A1:C1")
ct = cc["A1"]; ct.value = "Cadence, Channels & Production Notes"
ct.font = Font(name=FONT, size=14, bold=True, color=WHITE); ct.fill = PatternFill("solid", fgColor=INK)
ct.alignment = Alignment(horizontal="left", vertical="center", indent=1); cc.row_dimensions[1].height = 28

notes = [
 ("Rhythm","6 posts/week + 1 on-brand rest day","Posting on rest days would contradict the brand. Sun = story/poll only."),
 ("Primary platform","TikTok","Highest organic reach for cold consumer apps; lead here, cross-post everywhere."),
 ("Cross-post","IG Reels + YouTube Shorts","Same master edit; use Adobe-style resize per aspect ratio (9:16 master)."),
 ("Best post times","12–1pm & 7–9pm local","Lunch + evening doom-scroll windows; test and let data move them."),
 ("Master format","9:16, 7–15s, hard cut in <1s","No logo intro, no slow build. Hook frame must be loud on frame 1."),
 ("Captions","Always burned-in","Most watch muted — verbal hook ≈ the text overlay."),
 ("Loop","Make the spin loop seamlessly","Boosts watch-time & replays — the algorithm's favourite signal."),
 ("CTA","One soft CTA","'link in bio' or 'comment WHEEL'. Never hard-sell."),
 ("Series to milk","#11/#18 (wheel-runs-my-day), #16 (couples chores)","Franchises > singles. Number the episodes."),
 ("Sound","Wheel tick + completion chime","gen-audio asset (also fills the app's unimplemented wheelSoundEnabled)."),
 ("Roadmap gate","Body-double = 'soon', never 'now'","Not wired into the app yet — tease only (Day 30)."),
]
for c, h in enumerate(["Lever", "Setting", "Why / note"], 1):
    cell = cc.cell(row=2, column=c, value=h)
    cell.font = Font(name=FONT, size=10, bold=True, color=WHITE); cell.fill = PatternFill("solid", fgColor=CORAL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = border
cc.row_dimensions[2].height = 24
for i, n in enumerate(notes):
    for c, val in enumerate(n, 1):
        cell = cc.cell(row=3+i, column=c, value=val); cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.font = Font(name=FONT, size=9, color=INK)
        if c == 1: cell.font = Font(name=FONT, size=9, bold=True, color=CORAL)
        if (i % 2) == 1:
            cell.fill = PatternFill("solid", fgColor=BEIGE_LT)
for col, w in zip("ABC", [20,34,55]):
    cc.column_dimensions[col].width = w

# ----------------------------------------------------------------------------
# Sheet 5 — Summary (formulas)
# ----------------------------------------------------------------------------
sm = wb.create_sheet("Summary")
sm.merge_cells("A1:B1")
st = sm["A1"]; st.value = "Mix Check (auto-counts from the calendar)"
st.font = Font(name=FONT, size=14, bold=True, color=WHITE); st.fill = PatternFill("solid", fgColor=INK)
st.alignment = Alignment(horizontal="left", vertical="center", indent=1); sm.row_dimensions[1].height = 28

CALRANGE_PILLAR = "'30-Day Calendar'!$E$4:$E$33"
CALRANGE_PHASE = "'30-Day Calendar'!$C$4:$C$33"

def block(title, row):
    cell = sm.cell(row=row, column=1, value=title)
    cell.font = Font(name=FONT, size=10, bold=True, color=WHITE); cell.fill = PatternFill("solid", fgColor=CORAL)
    sm.cell(row=row, column=2).fill = PatternFill("solid", fgColor=CORAL)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

block("Posts by pillar", 3)
pill_labels = ["Relatable pain","Differentiator: kind+rest","Oddly satisfying","Niche","Trend-jack","Proof & momentum","Rest day (on-brand)"]
# Use wildcard COUNTIF since pillar cells carry sub-labels (e.g. 'Niche · ADHD')
pill_match = ["Relatable pain","Differentiator*","Oddly satisfying","Niche*","Trend-jack","Proof*","Rest day*"]
r = 4
for lab, m in zip(pill_labels, pill_match):
    sm.cell(row=r, column=1, value=lab).font = Font(name=FONT, size=9, color=INK)
    f = sm.cell(row=r, column=2, value=f'=COUNTIF({CALRANGE_PILLAR},"{m}")')
    f.font = Font(name=FONT, size=9, color=INK); f.alignment = Alignment(horizontal="center")
    r += 1
sm.cell(row=r, column=1, value="Total rows").font = Font(name=FONT, size=9, bold=True, color=INK)
tt2 = sm.cell(row=r, column=2, value=f'=COUNTA({CALRANGE_PILLAR})')
tt2.font = Font(name=FONT, size=9, bold=True, color=INK); tt2.alignment = Alignment(horizontal="center")

block("Posts by phase", r+2)
ph_labels = ["Pre-launch","LAUNCH WEEK","Use-cases","Retention"]
rr = r+3
for lab in ph_labels:
    sm.cell(row=rr, column=1, value=lab).font = Font(name=FONT, size=9, color=INK)
    f = sm.cell(row=rr, column=2, value=f'=COUNTIF({CALRANGE_PHASE},"{lab}")')
    f.font = Font(name=FONT, size=9, color=INK); f.alignment = Alignment(horizontal="center")
    rr += 1

block("Posting vs rest", rr+1)
sm.cell(row=rr+2, column=1, value="Posting days").font = Font(name=FONT, size=9, color=INK)
f1 = sm.cell(row=rr+2, column=2, value=f'=COUNTA({CALRANGE_PILLAR})-COUNTIF({CALRANGE_PILLAR},"Rest day*")')
f1.font = Font(name=FONT, size=9, color=INK); f1.alignment = Alignment(horizontal="center")
sm.cell(row=rr+3, column=1, value="Rest days").font = Font(name=FONT, size=9, color=INK)
f2 = sm.cell(row=rr+3, column=2, value=f'=COUNTIF({CALRANGE_PILLAR},"Rest day*")')
f2.font = Font(name=FONT, size=9, color=INK); f2.alignment = Alignment(horizontal="center")
sm.column_dimensions["A"].width = 28
sm.column_dimensions["B"].width = 12

import os
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "content-calendar-30day.xlsx")
wb.save(out)
print("saved:", out)
